# ExcelファイルのA列に記載された日本語を翻訳し、B列に出力するスクリプト。 1行目はヘッダーとして扱う。
# 使い方: python excel_translator.py <xlsx_file> [<lang_name>]
# <xlsx_file>: 入力・出力するExcelファイル
# <lang_name>: 言語名
#   - 翻訳先の言語名(任意。設定しない場合はB列のヘッダー部分に記載された言語名を使用)

import sys
import re
import gemma2_accessor
import openpyxl
import logging

# ログの設定
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')

class ExcelTranslator:
    def __init__(self, xlsx_file, lang_name=None):
        self.xlsx_file = xlsx_file
        self.lang_name = lang_name
        self.gemma2_accessor = gemma2_accessor.Gemma2Accessor()
        logging.info(f"ExcelTranslator initialized with file: {xlsx_file} and language: {lang_name}")

    def translate(self):
        logging.info("Starting translation process")
        wb = openpyxl.load_workbook(self.xlsx_file)
        sht = wb.active
        logging.debug("Excel file opened successfully")

        if self.lang_name is None:
            self.lang_name = sht['B1'].value
            if not self.lang_name:
                self.lang_name = "英語"
            logging.info(f"Translation language set to: {self.lang_name}")

        for row_idx, row in enumerate(sht.iter_rows(min_row=2, max_col=1, values_only=True), start=2):
            if row[0] is None:
                continue
            text = row[0]
            logging.debug(f"Translating text: {text}")

            translated_text = self.gemma2_accessor.translate(text, lang_name=self.lang_name)
            logging.debug(f"Translated text: {translated_text}")

            translated_text = re.sub(r"\n\s*\n.*", "", translated_text, flags=re.DOTALL).split("\n")[0]
            logging.debug(f"Processed translated text: {translated_text}")

            sht.cell(row=row_idx, column=2, value=translated_text)
            logging.debug(f"Written translated text to cell: B{row_idx}")

        wb.save(self.xlsx_file)
        logging.info("Translation process completed and file saved")

def main():
    if len(sys.argv) < 2:
        print("Usage: python excel_translator.py <xlsx_file> <lang_name>")
        print("<xlsx_file>: 入力・出力するExcelファイル")
        print("<target_lang>: 翻訳先の言語名(任意。設定しない場合はB列のヘッダー部分に記載された言語名を使用)")
        print('　※日本語の場合は、"日本語" とか "Japanese" とか、英語の場合は "英語" とか "English" とか')
        sys.exit(1)
    xlsx_file = sys.argv[1]
    lang_name = None
    if len(sys.argv) >= 3:
        lang_name = sys.argv[2]
    logging.info(f"Script started with file: {xlsx_file} and language: {lang_name}")
    excel_translator = ExcelTranslator(xlsx_file, lang_name)
    excel_translator.translate()

if __name__ == "__main__":
    main()
